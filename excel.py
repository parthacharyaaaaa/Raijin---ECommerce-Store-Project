import openpyxl
def addToExcel(finalPrice, time, cart, email, phoneNum) -> None:
    print("Excel called")
    print(finalPrice, time, cart, email, phoneNum)

    database = openpyxl.load_workbook('instance/orders.xlsx')
    sheet = database.active

    #adding entries to excel sheet
    emptyRow = 1
    for cell in sheet.iter_rows(values_only = True):
        if cell == "":
            break
        else:
            emptyRow = emptyRow + 1
    print(emptyRow)
    sheet.cell(row=emptyRow, column = 1, value=emptyRow-1)
    sheet.cell(row=emptyRow, column = 2, value=time)
    sheet.cell(row=emptyRow, column = 3, value=phoneNum)
    sheet.cell(row=emptyRow, column = 4, value=email)
    product = list(cart.keys())
    quantities = list(cart.values())
    for emptyCells in range(len(cart)):
        sheet.cell(row = emptyRow, column = emptyCells+5, value=f"{product[emptyCells]} - {quantities[emptyCells]}")
    
    database.save("instance/orders.xlsx")

def addContact(*args) -> None:
    print("Called addContact")
    database = openpyxl.load_workbook('instance/contactSubmissions.xlsx')
    sheet = database.active

    #adding entries to excel sheet
    emptyRow = 1
    for cell in sheet.iter_rows(values_only = True):
        if cell == "":
            break
        else:
            emptyRow = emptyRow + 1
    print(emptyRow)

    print(args)
    for emptyCells in range(len(args)):
        sheet.cell(row = emptyRow, column = emptyCells+1, value=args[emptyCells])
        print(args[emptyCells])

    database.save("instance/contactSubmissions.xlsx")